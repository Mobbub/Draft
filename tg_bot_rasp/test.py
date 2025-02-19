# # import datetime

# # def get_previous_monday(date_str):
# #   """
# #   Возвращает дату предыдущего понедельника для заданной даты в формате дд.мм.гггг.

# #   Args:
# #     date_str: Строка с датой в формате дд.мм.гггг.

# #   Returns:
# #     Строка с датой предыдущего понедельника в формате дд.мм.гггг.
# #     Возвращает None, если date_str не соответствует формату.
# #   """
# #   try:
# #     date_obj = datetime.datetime.strptime(date_str, "%d.%m.%Y").date()
# #     weekday = date_obj.weekday()  # 0 - понедельник, 6 - воскресенье
# #     days_to_monday = (weekday + 7) % 7 # Вычисляем сколько дней нужно отнять до понедельника
# #     previous_monday = date_obj - datetime.timedelta(days=days_to_monday)
# #     return previous_monday.strftime("%d.%m.%Y")
# #   except ValueError:
# #       return None


# # # Пример использования
# # date_str = "19.12.2024"
# # previous_monday = get_previous_monday(date_str)

# # if previous_monday:
# #     print(f"Для даты {date_str}, предыдущий понедельник: {previous_monday}")
# # else:
# #     print("Неверный формат даты.")

# import datetime

# # Получаем текущую дату
# today = datetime.datetime.now()

# # Получаем номер дня недели (0 - понедельник, 6 - воскресенье)
# day_number = today.weekday()

# # Словарь с названиями дней недели на русском
# days_in_russian = {
#     0: "Понедельник",
#     1: "Вторник",
#     2: "Среда",
#     3: "Четверг",
#     4: "Пятница",
#     5: "Суббота",
#     6: "Воскресенье"
# }

# # Получаем название дня недели
# day_of_week = days_in_russian[day_number]

# # Печатаем день недели
# print("Сегодня день недели:", day_of_week)


# import openpyxl

# # Замените 'your_file.xlsx' на путь к вашему файлу
# file_path = 'shedules/02.01.2025_downloaded_file.xlsx'

# # Загружаем файл
# workbook = openpyxl.load_workbook(file_path)

# # Получаем второй лист (индексация начинается с 0, поэтому второй лист - это 1)
# sheet = workbook.worksheets[2]

# # Перебираем все строки в листе и выводим данные
# for row in sheet.iter_rows(values_only=True):
#     print(row)

import openpyxl

data = []

# # Замените 'your_file.xlsx' на путь к вашему файлу
file_path = 'shedules/23.01.2025_DownloadedFile.xlsx'

# # Загружаем файл
workbook = openpyxl.load_workbook(file_path)

# # Получаем второй лист
sheet = workbook.worksheets[2]

# cell_address = 'A1'  # Замените на необходимый адрес ячейки

# # Получаем значение из указанной ячейки
# cell_value = sheet[cell_address].value

# # Выводим результат
# print(f'Значение в ячейке {cell_address}: {cell_value}')

# import re
# from datetime import datetime

# # Ваша строка
# input_string = "Четная неделя: (с 09.09; 23.09; 07.10; 21.10; 04.11;18.11; 02.12; 16.12)"

# # Используем регулярное выражение для поиска дат
# dates = re.findall(r'\d{2}\.\d{2}', input_string)

# for date in dates:
#     day, month = date.split('.')
#     year = '2023'  # здесь можно указать нужный год
#     formatted_date = f"{day}.{month}.{year}"
#     print(formatted_date)




# Перебираем все строки в листе
for row in sheet.iter_rows():
    # Создаем новый список для непустых значений и их адресов
    non_none_values_with_coords = [
        (cell.coordinate, cell.value) for cell in row if cell.value is not None
    ]

    # Если в строке есть непустые значения, выводим их с адресами
    if non_none_values_with_coords:
        data.append(non_none_values_with_coords)

parsed_data = {}

# Проходим по каждому набору данных
for row in data:
    current_row = {}
    
    for cell in row:
        coord, value = cell
        # Сохраняем значение по ключу координаты
        current_row[coord] = value
    
    # Добавляем текущую строку в общий словарь по индексу строки
    parsed_data[len(parsed_data) + 1] = current_row

# Выводим распарсенные данные
for row_number, values in parsed_data.items():
    print(f"Row {row_number}: {values}")