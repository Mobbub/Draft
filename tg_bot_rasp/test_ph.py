# from openpyxl import load_workbook
# from PIL import Image, ImageDraw, ImageFont

# def create_excel_image(data, cell_padding=10, font_size=20, exclude_texts=None):
#     # Загрузка шрифта, поддерживающего кириллицу
#     try:
#         font = ImageFont.truetype('arial.ttf', font_size)
#     except IOError:
#         print("Не удалось загрузить шрифт. Используется стандартный шрифт.")
#         font = ImageFont.load_default()

#     # Определяем размеры ячеек
#     cell_widths = []
#     cell_heights = []

#     # Сначала определяем размеры в зависимости от содержания
#     for row_idx, row in enumerate(data):
#         for col_idx, cell in enumerate(row):
#             # Проверяем, нужно ли игнорировать ячейку
#             if exclude_texts and cell in exclude_texts:
#                 continue

#             if cell is not None:
#                 # Создаем временное изображение для измерения текста
#                 temp_image = Image.new('RGB', (1, 1))  # Минимальный размер
#                 temp_draw = ImageDraw.Draw(temp_image)

#                 # Измеряем размеры текста
#                 bbox = temp_draw.textbbox((0, 0), str(cell), font=font)
#                 text_width = bbox[2] - bbox[0] + 2 * cell_padding
#                 text_height = bbox[3] - bbox[1] + 2 * cell_padding

#                 # Обновляем максимальную ширину колонки
#                 if col_idx >= len(cell_widths):
#                     cell_widths.append(text_width)
#                 else:
#                     cell_widths[col_idx] = max(cell_widths[col_idx], text_width)

#                 # Обновляем максимальную высоту строки
#                 if row_idx >= len(cell_heights):
#                     cell_heights.append(text_height)
#                 else:
#                     cell_heights[row_idx] = max(cell_heights[row_idx], text_height)

#     # Определяем размеры изображения
#     width = sum(cell_widths)
#     height = sum(cell_heights)

#     # Создаем новое изображение с белым фоном
#     image = Image.new('RGB', (width, height), 'white')
#     draw = ImageDraw.Draw(image)  # Создаем объект draw после создания изображения

#     # Рисуем ячейки и текст
#     for row_idx, row in enumerate(data):
#         for col_idx, cell in enumerate(row):
#             # Проверяем, нужно ли игнорировать ячейку
#             if exclude_texts and cell in exclude_texts:
#                 continue

#             if cell is not None:
#                 # Вычисляем координаты ячейки
#                 x0 = sum(cell_widths[:col_idx])
#                 y0 = sum(cell_heights[:row_idx])
#                 x1 = x0 + cell_widths[col_idx]
#                 y1 = y0 + cell_heights[row_idx]

#                 # Рисуем прямоугольник ячейки
#                 draw.rectangle([x0, y0, x1, y1], outline='black', fill='lightgray')

#                 # Добавляем текст в ячейку
#                 draw.text((x0 + cell_padding, y0 + cell_padding), str(cell), fill='black', font=font)

#     return image

# def load_data_from_excel(file_path):
#     workbook = load_workbook(file_path)
#     sheet = workbook.worksheets[2]
#     data = []

#     for row in sheet.iter_rows(values_only=True):
#         filtered_row = [cell for cell in row if cell is not None]  # Фильтруем None значения
#         if filtered_row:  # Добавляем только непустые строки
#             data.append(filtered_row)

#     return data

# # Убедитесь, что путь указан правильно
# excel_file_path = 'shedules/23.01.2025_DownloadedFile.xlsx'

# # Загрузка данных

# data = load_data_from_excel(excel_file_path)

# # Укажите тексты, которые нужно исключить
# exclude_texts = ['Дни недели', 'пара', 'вид занятий', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 'вид занятий ']  # Замените на ваши тексты

# # Создаем изображение, если есть данные
# if data:
#     image = create_excel_image(data, exclude_texts=exclude_texts)
#     # Сохраняем изображение

#     image.save('excel_table_image.png', 'PNG')
# else:
#     print("Нет данных для визуализации.")

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

def create_excel_image(data, cell_padding=10, font_size=20, exclude_texts=None):
    # Загрузка шрифта
    try:
        font = ImageFont.truetype('arial.ttf', font_size)
    except IOError:
        print("Не удалось загрузить шрифт. Используется стандартный шрифт.")
        font = ImageFont.load_default()

    # Определяем размеры ячеек
    cell_widths = []
    cell_heights = []

    # Сначала определяем размеры в зависимости от содержания
    for row_idx, row in enumerate(data):
        for col_idx, cell in enumerate(row):
            # Проверяем, нужно ли игнорировать ячейку
            if exclude_texts and cell in exclude_texts:
                continue

            if cell is not None:
                # Создаем временное изображение для измерения текста
                temp_image = Image.new('RGB', (1, 1))  # Минимальный размер
                temp_draw = ImageDraw.Draw(temp_image)

                # Измеряем размеры текста
                bbox = temp_draw.textbbox((0, 0), str(cell), font=font)
                text_width = bbox[2] - bbox[0] + 2 * cell_padding
                text_height = bbox[3] - bbox[1] + 2 * cell_padding

                # Обновляем максимальную ширину колонки
                if col_idx >= len(cell_widths):
                    cell_widths.append(text_width)
                else:
                    cell_widths[col_idx] = max(cell_widths[col_idx], text_width)

                # Обновляем максимальную высоту строки
                if row_idx >= len(cell_heights):
                    cell_heights.append(text_height)
                else:
                    cell_heights[row_idx] = max(cell_heights[row_idx], text_height)

    # Определяем размеры изображения
    width = sum(cell_widths)
    height = sum(cell_heights)

    # Создаем новое изображение с белым фоном
    image = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(image)

    # Рисуем ячейки и текст
    for row_idx, row in enumerate(data):
        current_x = 0  # Начальная координата по x для текущей строки
        for col_idx, cell in enumerate(row):
            # Проверяем, нужно ли игнорировать ячейку
            if exclude_texts and cell in exclude_texts:
                continue

            if cell is not None:
                # Вычисляем координаты ячейки
                x0 = current_x
                y0 = sum(cell_heights[:row_idx])
                x1 = x0 + cell_widths[col_idx]
                y1 = y0 + cell_heights[row_idx]

                # Рисуем прямоугольник ячейки
                draw.rectangle([x0, y0, x1, y1], outline='black', fill='lightgray')

                # Добавляем текст в ячейку
                draw.text((x0 + cell_padding, y0 + cell_padding), str(cell), fill='black', font=font)

                # Обновляем текущую координату для следующей ячейки
                current_x += cell_widths[col_idx]

    return image

def load_data_from_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]  # Индекс листа начинается с 0
    data = []

    for row in sheet.iter_rows(values_only=True):

        filtered_row = [cell for cell in row if cell is not None]  # Фильтруем None значения
        if filtered_row:  # Добавляем только непустые строки
            data.append(filtered_row)

    return data

# Убедитесь, что путь указан правильно
excel_file_path = 'shedules/23.01.2025_DownloadedFile.xlsx'

# Загрузка данных
data = load_data_from_excel(excel_file_path)

# Укажите тексты, которые нужно исключить
exclude_texts = ['Дни недели', 'пара', 'вид занятий', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 'вид занятий ']  # Замените на ваши тексты

# Создаем изображение, если есть данные
if data:
    image = create_excel_image(data, exclude_texts=exclude_texts)
    # Сохраняем изображение
    image.save('excel_table_image.png', 'PNG')
else:
    print("Нет данных для визуализации.")
